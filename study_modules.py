#'''
#Class definitions for 'Study' Class
#encompasses:
#    openpyxl and pydicom objects
#    simple variables (like cell and column definitions)
#    methods:
#'''

# PEP 8 compliance is a work in progress... sigh 


import pydicom
from dicom_dictionary import DicomDictionary
import openpyxl
import random
import getpass
import os
from datetime import date, time, datetime, timedelta
import numpy as np  # For DICOM pixel operations


# ----------------------> Constants <-------------------------------

alphalistboth = ['a','b','c','d','e','f','g','h','i','j','k','l','m','o','n',
                 'p','q','r','s','t','u','v','w','x','y','z','A','B','C','D',
                 'E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S',
                 'T','U','V','W','X','Y','Z'
                ]
alphalistupper = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N',
                  'O','P','Q','R','S','T','U','V','W','X','Y','Z'
                 ]
alphalistlower = ['a','b','c','d','e','f','g','h','i','j','k','l','m','o',
                  'n','p','q','r','s','t','u','v','w','x','y','z' 
                 ]

digitlist = [ '0','1','2','3','4','5','6','7','8','9' ]

class FileStats_Class( ):
    # Basic counters etc. for use in file stats.
    # Keep it neat and easily dealt with by helper functions.
    
    # Stats for each iterated 'root' directory
    # Needs to be reset after each root is completed.
    dir_count = 0
    file_count = 0
    valid_DCM_file = 0
    nondicom = 0
    copyOK = 0
    copyfailed = 0
    anonok = 0
    anonfailed = 0
    ignored = 0

    # Overall stats for the whole operation
    all_dir_count = 0
    all_file_count = 0
    all_valid_DCM_file = 0
    all_copyOK = 0
    all_copyfailed = 0
    all_anonok = 0
    all_anonfailed = 0
    skipped_dcm_filenames = []
    not_DCM_filenames = []
    tag_warning = []

    def reset_sub( self ):
        # All values back to zero (initialised values)
        self.dir_count = 0
        self.file_count = 0
        self.valid_DCM_file = 0
        self.copyOK = 0
        self.copyfailed = 0
        self.anonok = 0
        self.anonfailed = 0
        self.nondicom = 0
        self.ignored = 0

    
    def start_subdir( self, subdirlist, filelist ):
        self.dir_count  += len( subdirlist )
        self.file_count += len( filelist )

    def subdir_complete( self ):
        # Nothing to do right now... ?Delete?
        return
    
    def update_rootdir_complete( self ):
        self.all_dir_count += self.dir_count
        self.all_file_count += self.file_count
        self.all_valid_DCM_file += self.valid_DCM_file
        self.all_copyOK += self.copyOK
        self.all_copyfailed += self.copyfailed
        self.all_anonok += self.anonok
        self.all_anonfailed += self.anonfailed


class BasicDetails():
    AnonName = ""
    AnonID = ""
    PatientID = ""
    StudyInstanceUID = ""
    delta = None

    testfilename = ""
    savefilename = ""
    def clean(self):
        self.AnonName = ""
        self.AnonID = ""
        self.PatientID = ""
        self.StudyInstanceUID = ""
        self.delta = None
        self.testfilename = ""
        self.savefilename = ""


class Study_Class( ):
    # Some constants to use in this library
    # These are column or cell references prefixed by the sheet they belong in
    # The constants could be moved into a separate 'from _ import *' as they
    # don't, need to be accessed through the class -I just want them global
    CurrStudy = BasicDetails()

    QUIET = False
    VERBOSE = False
    DEBUG = False
    NEW_PREAMBLE = b'deidentified by Batch DICOM DeIdentifier' + b'\x00' * 88

    XLSPAGE_TITLE = 'A1'

    XLSFRONT_STUDYTITLE_CELL = 'B2'
    XLSFRONT_PI_CELL = 'B3'
    XLSFRONT_IRB_CODE_CELL = 'B4'
    XLSFRONT_NUMBER_OF_STUDYIDS_CELL = 'B5' 

    XLSDATA_STUDYIDS = 'B'
    XLSDATA_DATEADDED = 'C'
    XLSDATA_TIMEADDED = 'D'
    XLSDATA_PATIENTNAME = 'G'
    XLSDATA_PATIENTID = 'I'
    XLSDATA_ACCESSIONNUMBER = 'K'
    XLSDATA_STUDYDATE = 'M'
    XLSDATA_STUDYTIME = 'N'
    XLSDATA_DT_DELTA = 'O'
    XLSDATA_STUDYUID = 'P'
    XLSDATA_STUDYDESCRIPTION = 'R'

    XLSLOG_DATE = 'B'
    XLSLOG_TIME = 'C'
    XLSLOG_ACTIVITY = 'E'
    XLSLOG_USER = 'G'
    XLSLOG_COMPUTER = 'H'

    Config_Start_Row = 3

    VR_Action_Col = 'A'
    VR_Val_Col = 'B'
    VR_RVal_Col = 'C'

    FLAG_COL = 'E'
    FLAG_VAL_COL = 'F'

    Tag_Name_Col = 'H'
    Tag_Group_Col = 'I'
    Tag_Element_Col = 'J'
    Tag_Action_Col = 'K'
    Tag_RVal_Col = 'L'

    vr_actions = {}
    vr_action_list = []
    list_of_vr_actions = [
        'set_all_AE', 'set_all_AS',
        'set_all_AT', 'set_all_CS',
        'set_all_DA', 'set_all_DS',
        'set_all_DT', 'set_all_FL',
        'set_all_FD', 'set_all_IS',
        'set_all_LO', 'set_all_LT',
        'set_all_OB', 'set_all_OD',
        'set_all_OF', 'set_all_OW',
        'set_all_PN', 'set_all_SH',
        'set_all_SL', 'set_all_SQ',
        'set_all_SS', 'set_all_ST',
        'set_all_TM', 'set_all_UI',
        'set_all_UL', 'set_all_UN',
        'set_all_US', 'set_all_UT'
        ]
    tag_actions = {}
    tag_action_list = []

    #Please do not mess with the _CELL definitions below-
    # They need to tie in with the code to generate the
    # new blank XLS templates in create_new_study()
    # not sure how else to do it
    flag_list = [
        'DEL_PRIVATE_FLAG',
        'DEL_CURVES_FLAG',
        'CROP_US_TOPBAR_FLAG',
        'SET_DATE',
        'SET_TIME',
        'PRESERVE_PT_TIMELINE'
        ]
    flag_default = [
        'TRUE',  # Do delete private flags
        'TRUE',  # Do delete Curves
        '65px',  # Crop topbar in US image by 65
        '20101010',  # default date
        '120000.00',  # default time
        'TRUE'  # Maintain timeline for each patient by using a datetime delta
        ]
    flag_dict = {}  # index will be the flag name, contains set value from XLS
    flag_cell = {}

    
    dt_delta = {}  #  Date Time Delta dictionary.

    # This may now be obsolete so long as flag_list order is maintained.
    #DEL_PRIVATE_FLAG_CELL = FLAGS_COL + '3'
    #DEL_CURVES_FLAG_CELL = FLAGS_COL + '4'
    #CROP_US_TOPBAR_FLAG_CELL = FLAGS_COL + '5'

    xls_UID_lookup             = {}
    test_study_UID             = ''
    
    next_log_row               = 0
    next_studyID_row           = 0

    # These will have values assigned after the workbook is opened
    frontsheet = None
    datasheet = None
    logsheet = None

    # log levels: debug = 3, high = 2, normal = 1
    # This should reflect the default, then over-ridden by cmd line options.
    LOGLEVEL_NORMAL = 1
    LOGLEVEL_HIGH = 2
    LOGLEVEL_DEBUG = 3
    GLOBAL_LOGLEVEL = 1
    LOGLEVEL_TXT = {LOGLEVEL_NORMAL: '',
                    LOGLEVEL_HIGH: 'High', 
                    LOGLEVEL_DEBUG: 'DEBUG'
                   }

    # Get these only once at the start of runtime.
    log_username = getpass.getuser()
    log_computername = os.environ['COMPUTERNAME']

    # List of filenames to ignore. These will be skipped and not copied.
    SKIP_LIST = ['DICOMDIR',
                 'VERSION',
                 'LOCKFILE'
                ]

    # List of tags to raise warnings for if they are not present
    # in the loaded DCM files.
    DCM_TAG_CHECKLIST = ['PatientID',
                         'PatientName',
                         'AccessionNumber',
                         'StudyInstanceUID',
                         'StudyDate'
                        ]

    # Assign the pydicom and openpyxl objects at start.
    # Some pylint warning were raised when DCM was set to 'False'
    DCM = pydicom.Dataset()
    XLS = openpyxl.Workbook()

    # DateTime variables
    # Default reference date is 15th July 2015, midday
    # This should be configured by XLS
    REF_YEAR = 2015
    REF_MONTH = 7
    REF_DAY = 15
    REF_HH = 12
    REF_MM = 00
    REF_SS = 00


    # *********************************************************************
    # *                                                                   *
    # *                     Start of Methods                              *
    # *                                                                   *
    # *                                                                   *
    # *********************************************************************

    def __init__(self, QUIET = False, VERBOSE = False, DEBUG = False ):
        # Duplicate some flags from ops class
        if QUIET:
            self.QUIET = True
        if VERBOSE:
            self.VERBOSE = True
        # Set global log level
        if DEBUG:
            self.DEBUG = True
            self.GLOBAL_LOGLEVEL = self.LOGLEVEL_DEBUG
            self.msg(f'Logging set to {self.LOGLEVEL_TXT[self.GLOBAL_LOGLEVEL]}', level='VERBOSE')
        else:
            self.GLOBAL_LOGLEVEL = self.LOGLEVEL_NORMAL
        
        # Load flag defaults into flag_dict dict.
        # enter flag cell location into flag_cell dict.
        line = 0
        for item in self.flag_list:
            self.flag_dict[ item ] = self.flag_default[line]
            self.flag_cell[ item ] = self.FLAG_VAL_COL + str( line + self.Config_Start_Row )
            line += 1
        for item in self.flag_dict:
            print(f'__init__ flag_dict[{item}] = {self.flag_dict[item]}')


    def _update_fromDCM(self):
        #self.CurrStudy.AnonName = ""  # Not yet created
        #self.CurrStudy.AnonID = ""  # Not yet created
        self.CurrStudy.PatientID = self.DCM.PatientID
        self.CurrStudy.StudyInstanceUID = self.DCM.StudyInstanceUID
        # delta is cached on loading XLS in study.load_XLS
        if self.CurrStudy.PatientID in self.dt_delta:
            self.CurrStudy.delta = self.dt_delta[self.CurrStudy.PatientID]


    # ####################################>  Helper function for .load_xls()
    def import_xls_settings( self ):
        # setup standard human-readable shortcuts (more readable for me...)
        self.frontsheet = self.XLS['Front']
        self.datasheet = self.XLS['Data']
        self.logsheet = self.XLS['Log']
        self.cfgsheet = self.XLS['Config']

        # Identify the first available log row
        self.find_first_available_log_row()

        # Load the existing UIDs into the dict cache
        self.cache_existing_xls_UIDs( )
        self.cache_dt_delta()

        # Identify row of the first available new studyID
        self.next_studyID_row = self.first_available_studyID_row( )

        # Load config from the Config sheet <--------------------------
        # Flags:  to tidy- do I need both the FLAG and flags dict?
        self.readXLSFlags()
        self.displayXLSFlags()

        # Read VR based actions
        row = self.Config_Start_Row
        action = str(self.cfgsheet[ self.VR_Action_Col + str(row) ].value).lower()
        enabled = self.cfgsheet[ self.VR_Val_Col + str(row) ].value

        while action != None:
            if enabled:
                value_rep = action[-2:].upper()
                vr_action_list = [
                    action,
                    enabled,
                    self.cfgsheet[ self.VR_RVal_Col + str(row) ].value
                    ]
                self.vr_actions[value_rep] = vr_action_list
                print(f'row: {row}: {vr_action_list[-1]}')
            row += 1
            action = self.cfgsheet[ self.VR_Action_Col + str(row) ].value
            enabled = self.cfgsheet[ self.VR_Val_Col + str(row) ].value

        # Tag based actions
        row = self.Config_Start_Row
        rowstr = str(row)
        action = self.cfgsheet[ self.Tag_Action_Col + str(row) ].value

        while action != None:
            group = hex( int( self.cfgsheet[ self.Tag_Group_Col + rowstr].value, 16) )
            element = hex( int(self.cfgsheet[ self.Tag_Element_Col + rowstr].value, 16) )
            repvalue = str(self.cfgsheet[ self.Tag_RVal_Col + rowstr ].value)
            action = str(action)
            # Update tag name column in config sheet
            tagname = DicomDictionary.get( combotag( group, element ), [0,0,'** Unknown **'] )[2]  # [2] selects the tag name
            self.cfgsheet[ self.Tag_Name_Col + rowstr ].value = tagname
            self.tag_action_list.append([
                group,
                element,
                action,
                repvalue
                ])
            #print(f'row: {row}: ({group},{element}) action: {action}, rep value: {repvalue}')
            row += 1
            rowstr = str(row)
            action = self.cfgsheet[ self.Tag_Action_Col + rowstr ].value




    ####################################################> load_xls()
    def load_xls( self, xls_filename ):
        """
        Loads study XLS file, and does basic checking to make sure that it
        is actually a valid XLS sheet with check_xls_is_valid().
        Usage: study_workbook = load_study_xls( <filename> )
        Returns: openpyxl XLS workbook object
        """
        # mini-__init__
        self.xls_UID_lookup = {}  # new lookup dict for study UIDs 
        valid = True  # until proven false
        validity_msg = ""  # info about errors as we find them

        # Test for openpyxl failing to open XLS file
        try:
            self.XLS = openpyxl.load_workbook( xls_filename )
        except:
            print(f'load_xls: {xls_filename} failed to load. FATAL ERROR')
            raise

        # Test actually saving the XLS file before proceeding.
        try:
            self.XLS.save( xls_filename )
        except:
            print(f'{xls_filename} save permission denied.')
            print('Is the file open in excel?.\nPlease unlock and try again.')
            raise

        #-------------------   Validity Checks go here -----------

        if 'Front' not in self.XLS.sheetnames:
            valid = False
            validity_msg += "No \'Front\' sheet;"

        if 'Log' not in self.XLS.sheetnames:
            valid = False
            validity_msg += "No \'Log\' sheet;"

        if 'Data' not in self.XLS.sheetnames:
            valid = False
            validity_msg += "No \'Data\' sheet;"

        if 'Config' not in self.XLS.sheetnames:
            valid = False
            validity_msg += "No \'Config\' sheet;"

        #---------------------------------------------------------

        if valid:
            self.import_xls_settings(  )
            self.log( f'load_xls: Complete OK. loaded {xls_filename} OK',
                     self.LOGLEVEL_DEBUG
                    )
            return True
        else:
            print(f'load_xls: {xls_filename} failed checks. FATAL ERROR')
            return False


    #----------------------------------------------------------------------
    # Create new 'blank' XLS. Does not load a blank XLS to do this.

    def new_XLS( self ):
        self.XLS = openpyxl.Workbook()
        self.xls_populate_attribs()


    def xls_populate_attribs( self ):
        '''Used in Init_Study.py after creating a new [blank] XLS file object
        Populate things like shortcuts to the workbook sheets etc.
        This is essentially moved from the load_xls method.
        '''
        #Create readable shortcuts
        self.frontsheet = self.XLS['Front']
        self.datasheet = self.XLS['Data']
        self.logsheet = self.XLS['Log']
        self.cfgsheet = self.XLS['Config']
        
        #perform baseline evaluation of all new workbooks
        self.find_first_available_log_row()
        self.cache_existing_xls_UIDs( )
        self.cache_dt_delta()
        self.next_studyID_row = self.first_available_studyID_row( )

        self.log( f'xls_repopulate_attribs: Complete.', self.LOGLEVEL_DEBUG )


    def write_xls_to_disc(self, filename):
        '''Alternative route to the openpyxl .save method
        '''
        self.XLS.save( filename )


    def cache_existing_xls_UIDs(self):
        """Identifies each stored study UID and caches both it and its row number
        in a dictionary object.\n
        Returns the absolute row number - not relative
        -so starts at 2 (as does the data) and ends at total_studyIDs + 1\n
        Usage: <obj>.cache_existing_xls_UIDs()\n
        Returns:  True (no probs), False (error)  
        """
        self.log('.cache_existing_xls_UIDs(): Started.', self.LOGLEVEL_DEBUG )
        
        # Probably unnecessary re-definition/reset
        self.xls_UID_lookup = {}

        row = 2
        max_row = self.frontsheet[ self.XLSFRONT_NUMBER_OF_STUDYIDS_CELL ].value + 1
        # +1 as the data-containing rows start at row 2, not 1
        max_row = int( max_row )

        check_ptID = self.datasheet[self.XLSDATA_PATIENTID + str(row)].value
        check_studyID = self.datasheet[self.XLSDATA_STUDYIDS + str(row)].value
        check_study_UID = self.datasheet[self.XLSDATA_STUDYUID + str(row)].value
        #raw_dt_delta = None

        while (check_studyID != None ) and (check_ptID != None ) and (row <= max_row ):
            self.xls_UID_lookup[ check_study_UID ] = row
            #if raw_dt_delta:
            #    self.dt_delta[ check_ptID ] = self.delta_str2obj(raw_dt_delta)
            row += 1
            check_ptID      = self.datasheet[ self.XLSDATA_PATIENTID + str(row) ].value
            check_studyID   = self.datasheet[ self.XLSDATA_STUDYIDS  + str(row) ].value
            check_study_UID = self.datasheet[ self.XLSDATA_STUDYUID  + str(row) ].value
            #raw_dt_delta = self.datasheet[ self.XLSDATA_DT_DELTA + str(row)].value

        self.log( f'self.cache_existing_xls_UIDs: Completed OK. Found&cached {len(self.xls_UID_lookup)} existing UIDs.  Final row={row}', self.LOGLEVEL_HIGH )
        return True


    def cache_dt_delta( self ):
        '''Load all encountered original pt IDs and associated datetime delta
        Extract string from XLS.Data dt_delta column.
        Convert to datetime.timedelta object
        Store timedelta object in dt_delta dictionary
        Return: None
        Params: None
        '''
        #  self.XLSDATA_DATETIME_DELTA is the log column
        #  dt_delta is the cache dictionary.
        #  expected format "days:seconds.miliseconds"
        #  dt_delta key is real ptID (as same delta for that pt)
        self.msg('.cache_dt_delta: caching deltas from XLS')
        self.log('.cache_dt_delta: Started....', self.LOGLEVEL_DEBUG )
        row = 2
        max_row = self.frontsheet[ self.XLSFRONT_NUMBER_OF_STUDYIDS_CELL ].value
        max_row = int(max_row + 1)
        # +1 as the data-containing rows start at row 2, not 1

        ptID = self.datasheet[self.XLSDATA_PATIENTID + str(row)].value
        delta = None
        delta = self.datasheet[self.XLSDATA_DT_DELTA + str(row)].value

        while (ptID != None ):
            if (delta != None) and (ptID not in self.dt_delta):  
                # ONLY if a delta exists
                # AND ptID not already assigned a delta
                # delta is the raw XLS string value
                # convert into a timedelta object & store in dt_delta dict
                self.dt_delta[ ptID ] = delta_str2obj( delta )
            row += 1
            ptID = self.datasheet[self.XLSDATA_PATIENTID + str(row)].value
            delta = self.datasheet[self.XLSDATA_DT_DELTA + str(row)].value
            
        self.log( f'.cache_dt_delta(): Completed. Found&cached {len(self.dt_delta)} unique ptIDs & deltas.', self.LOGLEVEL_HIGH )
        self.msg(f'Found&cached {len(self.dt_delta)} unique ptIDs & deltas.')
        return True
    

    def find_first_available_log_row( self ):
        '''
        Identify the 1st free log row into which to write new messages\n
        Usage: Study_Class.find_first_available_log_row( )\n
        Returns: Int row number
        '''
        row = 2   # This is the first possible logging row

        log_check = self.logsheet[ self.XLSLOG_ACTIVITY + str(row) ].value
    
        # if there is an activity msg (ie != None) 
        # then go to the next row and check again
        while (log_check is not None ):
            row += 1
            log_check = self.logsheet[ self.XLSLOG_ACTIVITY + str(row) ].value

        self.next_log_row = row
        self.log(f'find_first_available_log_row: row={self.next_log_row}',
                 self.LOGLEVEL_DEBUG
                )

        return row
    

    def readXLSFlags( self ):

        for item in self.flag_list:
            self.flag_dict[item] = self.cfgsheet[ self.flag_cell[item] ].value
            #self.msg(f'Flag {item}: {self.flag_dict[item]} (new method)')
        
        for item in self.flag_dict:
            print(f'readXLSFlags: flag_dict[{item}] = {self.flag_dict[item]}')


    def displayXLSFlags( self ):
        '''displayFlags()
        Print to screen the flags and if enabled or not.
        Only prints if Normal or Verbose, not Quiet
        '''
        if self.QUIET:
            return
        for item in self.flag_list:
            self.msg(f'Flag {item.ljust(25," ")}: { self.flag_dict[item] }  ')



        if self.flag_dict['CROP_US_TOPBAR_FLAG']:
            self.msg('\tWarning: US image blanking does not support compressed images.  Output must be checked for residual PHI.')


    #------------------------------------------------------------------------------------------------------
    def log( self, message_str, msg_log_level=LOGLEVEL_NORMAL ):
        '''
        This is supposed to add a new line to the log page in the XLS file, describing a change.
        Perhaps this is a good place to set log level...
        log levels: debug = 3, high = 2, normal = 1
        '''
        # Only logs messages that are at or below the global log level. 
        # So debug msgs will bot be logged in high or normal logging
        if ( msg_log_level > self.GLOBAL_LOGLEVEL ):
            return False

        # log date, time, activity, user, computer
        row_text = str( self.next_log_row )
        self.logsheet[ self.XLSLOG_ACTIVITY + row_text ] = f'({ self.LOGLEVEL_TXT[ msg_log_level ] }): { message_str }'
        dateobject = datetime.now()

        self.logsheet[ self.XLSLOG_DATE     + row_text ] = dateobject.strftime('%d-%m-%Y') # date  
        self.logsheet[ self.XLSLOG_TIME     + row_text ] = dateobject.strftime('%H:%M:%S') # timenow
        self.logsheet[ self.XLSLOG_USER     + row_text ] = self.log_username               # username
        self.logsheet[ self.XLSLOG_COMPUTER + row_text ] = self.log_computername           # compname
        # Finally increment the Log row pointer
        self.next_log_row += 1
        return True
    

    def msg(self, message, level = 'NORMAL', endstr = '\n' ):
		# level should correlate with Quiet/Normal/Verbose string
		# if level arg is omitted, it will default to 'NORMAL'
        if self.QUIET:  # Quiet prints nothing
            return
        if self.VERBOSE or self.DEBUG:  # verbose prints everything
            print( message, end = endstr )
            return
        # Only not QUIET or VERBOSE reaches this point.
        if level == 'NORMAL':
            print( message, end = endstr )


    def get_DCM_StudyInstanceUID( self ):
        '''
        Returns the Study Instance UID from the specified DICOM file object
        Usage:   string_variable = <object>.get_DCM_StudyInstanceUID( )
        Returns:  a string containing the study instance UID.
                            returns FALSE if no StudyInstanceUID tag is found.
        This iterates through all tags until it finds the correct one.
            This bypasses the issue with DICOMDIR hiding it in a ?series?
            which is not visible to the standard dicom_object.StudyInstanceUID method.
        '''
        siUID = False
        for elem in self.DCM.iterall():
            if 'Study Instance UID' == elem.name:
                siUID = elem.value
                break   # Stops at 1st StudyInstanceUID
        # If no StudyInstanceUID tag is found, this returns False
        return siUID


    def get_old_study_attrib_from_UID( self, attribute, test_uid ):
        '''
        Returns string from relevent cell in Data sheet from the workbook.
        Takes the studyUID and str 'attribute' as reference
        Usage: get_old_study_attrib( <openpyxl workbook>, <existing study UID>, <str attribute> )
        'attribute' should be txt indicating the relevent data page column. 
        Use the static xls_Data_... column values from studytools.py
        eg. xls_Data_study_UID
        '''
        old_study_ID = self.datasheet[ attribute + str( self.xls_UID_lookup[ test_uid ] )  ].value
        return old_study_ID
    

    def first_available_studyID_row ( self ):
        """
        Identifies the FIRST available studyID in the XLS file
        This is slow, so separated from (but called by) the NEXT study ID generator
        Returns the absolute row number - not relative, so starts at 2 (as does the data) and ends at total_studyIDs + 1
        Usage: first_blank_ptID_row =  first_available_studyID_row ( <openpyxl xls workbook> )
        Returns: 1st free row where a studyID can be assigned. 
        """

        row = 2
        max_row = self.frontsheet[ self.XLSFRONT_NUMBER_OF_STUDYIDS_CELL ].value + 1  # +1 as the data-containing rows start at row 2, not 1

        check_ptID    = self.datasheet[ self.XLSDATA_PATIENTID + str(row)].value
        check_studyID = self.datasheet[ self.XLSDATA_STUDYIDS  + str(row)].value
    
        while (check_studyID != None ) and (check_ptID != None ) and (row <= max_row ):
            row += 1
            check_ptID    = self.datasheet[ self.XLSDATA_PATIENTID + str(row)].value
            check_studyID = self.datasheet[ self.XLSDATA_STUDYIDS  + str(row)].value

        return row



    def try_dcm_attrib( self, attrib_str, failure_value ):
        '''
        PYDICOM crashes if the queried data element is not present. Often happens in some anonymised DICOMs.
        This is a quick & dirty but 'safe' method to query and return something.
        Usage: some_var = <>.try_dcm_attrib( <attribute string>, <failure string> )
        Returns: on success returns the value held in the DICOM object attribute provided.
            On failure, returns the failure string.
        '''
        try:
                value = self.DCM.data_element( attrib_str ).value
        except:
                value = failure_value

        return value


    # - Works but maybe not as expected...
    def assign_next_free_studyID ( self ):   #XLS openpyxl workbook object
        '''
        Returns the new studyID, populates current DCM data into corresponding datasheet row & logs made
        and iterates down the studyID list with each new call.
        Built-in check to see if exceeded number of valid studyIDs
        
        '''
        self.log( 'running: assign_next_free_studyID()', self.LOGLEVEL_DEBUG )

        # If this is the 1st time running then do this
        if self.next_studyID_row == 0:
            self.next_studyID_row = self.first_available_studyID_row( )
        
        new_studyID    = self.datasheet[ self.XLSDATA_STUDYIDS + str( self.next_studyID_row ) ].value
        new_UID        = self.DCM.StudyInstanceUID
        no_of_studyIDs = self.frontsheet[ self.XLSFRONT_NUMBER_OF_STUDYIDS_CELL ].value

        # Check to see if we have exceeded available StudyIDs
        if self.next_studyID_row > ( no_of_studyIDs + 1) and new_studyID == None:
            self.log( '<obj>.assign_next_free_studyID: Run out of Study IDs in the xls file!!!', self.LOGLEVEL_NORMAL)
            return False
        elif self.next_studyID_row > ( no_of_studyIDs + 1) and new_studyID != None:
            # there is a mismatch- my code thinks we have exceeded the number of used studyIDs
            # However, new_studyID != None -ie the datacell in the XLS is not empty. We presume this cell contains a valid studyID...
            # A warning is logged however.
            self.log(f'<>.assign_next_free_studyID: WARNING -- next_studyID_row ({self.next_studyID_row}) is more than no_of_studyIDs ({no_of_studyIDs}) but XLScell is non-empty ({new_studyID}). Assuming it contains a valid studyID', self.LOGLEVEL_NORMAL)

        # Populate current pt data into XLS datasheet studyID into 
        dateobject = datetime.now()
        current_row_str = str( self.next_studyID_row )

        self.datasheet[ self.XLSDATA_DATEADDED        +  current_row_str ] = str( dateobject.strftime('%d-%m-%Y') )
        self.datasheet[ self.XLSDATA_TIMEADDED        +  current_row_str ] = str( dateobject.strftime('%H:%M:%S') )

        self.datasheet[ self.XLSDATA_PATIENTID        +  current_row_str ] = str( self.try_dcm_attrib( 'PatientID',        'Nil' ) )
        self.datasheet[ self.XLSDATA_PATIENTNAME      +  current_row_str ] = str( self.try_dcm_attrib( 'PatientName',      'Nil' ) )
    
        self.datasheet[ self.XLSDATA_ACCESSIONNUMBER  +  current_row_str ] = str( self.try_dcm_attrib( 'AccessionNumber',  'Nil' ) )
        self.datasheet[ self.XLSDATA_STUDYDATE        +  current_row_str ] = str( self.try_dcm_attrib( 'StudyDate',        'Nil' ) )
        self.datasheet[ self.XLSDATA_STUDYTIME        +  current_row_str ] = str( self.try_dcm_attrib( 'StudyTime',        'Nil' ) )
        self.datasheet[ self.XLSDATA_STUDYUID         +  current_row_str ] = str( self.try_dcm_attrib( 'StudyInstanceUID', 'Nil' ) )
        self.datasheet[ self.XLSDATA_STUDYDESCRIPTION +  current_row_str ] = str( self.try_dcm_attrib( 'StudyDescription', 'Nil' ) )

        #self.msg(f'orig date {self.DCM.StudyDate}  orig time {self.DCM.StudyTime}',self.LOGLEVEL_HIGH)
        self.CurrStudy.delta = self.create_dt_delta( self.DCM.StudyDate, self.DCM.StudyTime)
        self.datasheet[ self.XLSDATA_DT_DELTA + current_row_str ] = delta_obj2str( self.CurrStudy.delta )
        # DateTime Delta in 'days:seconds.microseconds' format

        # insert Logging message(s) here. Only 1 will be sent- depending on global_log_level
        if self.log( 'running: assign_next_free_studyID() in generator loop', self.LOGLEVEL_DEBUG ):
            pass
        elif self.log( f'Assigned { new_studyID } to { self.try_dcm_attrib(  "PatientID", "No_ID" ) }', self.LOGLEVEL_HIGH ):
            pass
        else:
            self.log( f'Assigned new StudyID to {new_UID}', self.LOGLEVEL_NORMAL )

        # Update the dict_cache.  Perhaps better to use in a class...
        self.xls_UID_lookup[ new_UID ] = self.next_studyID_row


        # Increment next_studyID_row to point to the next row
        self.next_studyID_row += 1

        # return the new studyID string  
        
        return new_studyID
    

    def deidentifyDICOM( self, newPtName = 'Anon', newPtID = 'research' ):
        '''Performs the removal of PHI
        Including applying the DateTime delta to ALL Date, Time and DateTime VRs
        '''
        self.msg('->deidentityDICOM...',level='DICOM')
        # enact FLAG based actions - need to sync this with XLS flag list
        if self.flag_dict['DEL_PRIVATE_FLAG']:
            self.DCM.remove_private_tags()
            self.msg('removing private tags', level='DEBUG')
        
        if self.flag_dict['DEL_CURVES_FLAG']:
            self.DCM.walk( del_curves_callback )
            self.msg('removig curves', level='DEBUG')

        if self.flag_dict['CROP_US_TOPBAR_FLAG'] and self.DCM.Modality=='US':
            self.blankTopBar()
            self.msg('blank topbar', level='DEBUG')

        # dt_delta:
        # if 'PRESERVE_PT_TIMELINE' is True, the delta will ne non-zero.
        # delta will be applied to all DA, TM and DT value representations.
        # delta is calculated on loading DICOM, in:
        # ->process_file -> study.assign_next_free_studyID()
        if self.DCM.PatientID in self.dt_delta:
            self.delta_apply_all_tags()
            self.msg('applying dt delta', level='DEBUG')

        # enact VR based actions
        self.DCM.walk(self.Perform_VR_Actions)

        # enact TAG-based actions
        for [group, element, action, repvalue] in self.tag_action_list:
            #print(f'{pretty_tag(group,element)} action: {action}, rep value: {repvalue}')
            if [group,element] in self.DCM:  # If the tag already exists
                exists = True
            else:
                exists = False
            
            if action.lower() == 'replace' and exists:  # ie If 'replace' and tag exists
                if [group,element] in self.DCM:
                    self.DCM[group, element].value = repvalue
            elif action.lower() == 'delete':
                if [group,element] in self.DCM:
                    del self.DCM[group,element]

        pass


    # callback to enact Value Representation (VR) based actions
    def Perform_VR_Actions(self, dataset, data_element):
        vr_str = data_element.VR
        if vr_str in self.vr_actions:
            new_val = self.vr_actions[vr_str][2]
            if new_val == None:
                new_val = ""
            data_element.value = new_val


    def create_dt_delta(self, oDate_raw:str, oTime_raw:str = "000000.000000")->timedelta:
        '''Take study date & time, and calculate delta with reference.
        Such that newDateTime = studyDateTime + delta

        Params: Take in original date & time of study as strings from XLS.
        NB- reference datetime is study_class.REF_YEAR/MONTH/DAY/HH/MM/SS/MS
        Return: DateTime object
        '''
        # delta should be in a format useable by the datetime module
        # DICOM date format is 'yyyymmdd', older format 'yyyy:mm:dd'
        # DICOM time format is 'HHMMSS.FFFFFF'
        # http://dicom.nema.org/dicom/2013/output/chtml/part05/sect_6.2.html
        oTime = remove( oTime_raw, ':.')  # strip out expected non-numerics
        oDate = remove( oDate_raw, ':.')  
        if not oTime:
            oTime = "000000000000"
        d_yyyy = int(oDate[:4])
        d_mm = int(oDate[4:6])
        d_dd = int(oDate[6:8])
        t_hh = int(oTime[0:2])
        t_mm = int(oTime[2:4])
        t_ss = int(oTime[4:6])
        if len(oTime[4:])>3:  # ie if the seconds contains float
            #  Get miliseconds
            t_ms = int( oTime[6:] )
        else:
            t_ms = 0

        oldDateTime = datetime( year=d_yyyy, month=d_mm, day=d_dd,
                                hour=t_hh, minute=t_mm, second=t_ss,
                                microsecond=t_ms
                                )
        refDateTime = datetime( year=self.REF_YEAR,
                        month=self.REF_MONTH,
                        day=self.REF_DAY,
                        hour=self.REF_HH,
                        minute=self.REF_MM,
                        second=self.REF_SS
                        )
        delta_obj = refDateTime - oldDateTime
        return delta_obj


    def delta_apply_all_tags(self ):
        '''Apply the known delta to ALL date/time/datetime tags
        Steps through all tags and changes all DA/TM/DT value representation tags
        Delta is zero if flag is not enabled.

        Params: None (just uses the Study_Class object)
        Returns: None
        '''
        self.msg(f' Applying delta...', endstr='')
        self.delta = self.dt_delta[ self.CurrStudy.PatientID ]
        # workaround to get delta into the callback
        # to which only the pyDICOM object is available.
        self.DCM.preamble = delta_obj2str(self.delta)
        # Step through all tags using the .walk pyDICOM method
        # use a callback to process tags.
        self.DCM.walk( delta_apply_callback )


    # ---------------------> Moved from Init_Study
    # should return the openpyxl workbook object 
    def create_new_study( 	self, 
                            new_xlsfilename,
                            new_study_title,
                            new_primary_investigator,
                            new_number_of_study_IDs ):


        #new_study = Study_Class()
        self.XLS = openpyxl.Workbook()    # create new blank XLS object
        
        try:
            self.XLS.save( new_xlsfilename ) # Sanity check to make sure we can save
        except:
            print(f'Fatal Error: Failed to save \"{ new_xlsfilename }\"')
            raise
        else:
            print(f'Created blank template file \"{ new_xlsfilename }\" OK')


        self.frontsheet = self.XLS.active
        self.frontsheet.title = 'Front'

        self.datasheet = self.XLS.create_sheet('Data') 
        self.logsheet = self.XLS.create_sheet('Log' ) 
        self.cfgsheet = self.XLS.create_sheet('Config')  

        #--------------------->  Add in basic data  <-------------------------
        # Col A is the list of titles. Text info is in col B.
        self.frontsheet['A1']                                                     = 'Front Page'            
        self.frontsheet['A' + str( self.XLSFRONT_STUDYTITLE_CELL[-1] ) ]         = 'Study Title:'          # data held in xlsFront_study_title_cell
        self.frontsheet['A' + str( self.XLSFRONT_PI_CELL[-1] )          ]         = 'Primary Investigator:' # data held in xlsFront_PI_cell
        self.frontsheet['A' + str( self.XLSFRONT_NUMBER_OF_STUDYIDS_CELL[-1] ) ] = 'No of Study IDs'       # data held in xlsFront_number_of_study_IDs_cell
        self.frontsheet['A' + str( self.XLSFRONT_IRB_CODE_CELL[-1] )    ]         = 'Study IRB Code'        # data held in xlsFront_IRB_code_cell

        self.frontsheet[ self.XLSFRONT_STUDYTITLE_CELL         ] = new_study_title
        self.frontsheet[ self.XLSFRONT_PI_CELL                  ] = new_primary_investigator
        self.frontsheet[ self.XLSFRONT_NUMBER_OF_STUDYIDS_CELL ] = new_number_of_study_IDs
        self.frontsheet[ self.XLSFRONT_IRB_CODE_CELL            ] = 'IRBcode1234'	

        # Row 1 is the column title row
        self.datasheet['A1']  = 'Data Page'
        self.datasheet[ self.XLSDATA_STUDYIDS + '1'] = 'Study IDs'
        self.datasheet[ self.XLSDATA_DATEADDED + '1'] = 'Date Added'
        self.datasheet[ self.XLSDATA_TIMEADDED + '1'] = 'Time Added'

        self.datasheet[ self.XLSDATA_PATIENTNAME + '1'] = 'Patient Name'

        self.datasheet[ self.XLSDATA_PATIENTID + '1'] = 'Patient ID'
        self.datasheet[ self.XLSDATA_ACCESSIONNUMBER + '1'] = 'Accession No.'
        self.datasheet[ self.XLSDATA_STUDYDATE + '1'] = 'Study Date'
        self.datasheet[ self.XLSDATA_STUDYTIME + '1'] = 'Study Time'
        self.datasheet[ self.XLSDATA_DT_DELTA + '1'] = 'DateTime Delta'
        self.datasheet[ self.XLSDATA_STUDYUID + '1'] = 'Study UID'
        self.datasheet[ self.XLSDATA_STUDYDESCRIPTION + '1'] = 'Study Description'

        self.logsheet['A1']   = 'Log Page'

        self.logsheet[ self.XLSLOG_DATE + '1'] = 'Date'
        self.logsheet[ self.XLSLOG_TIME + '1'] = 'Time'
        self.logsheet[ self.XLSLOG_ACTIVITY + '1'] = 'Log Activity'
        self.logsheet[ self.XLSLOG_USER + '1'] = 'User'
        self.logsheet[ self.XLSLOG_COMPUTER + '1'] = 'Computer'

        self.cfgsheet[ self.VR_Action_Col + '1' ] = 'ACTION'
        self.cfgsheet[ self.VR_Action_Col + '2' ] = '-'
        self.cfgsheet[ self.VR_Val_Col + '1' ] = 'Value'
        self.cfgsheet[ self.VR_Val_Col + '2' ] = '-'
        self.cfgsheet[ self.VR_RVal_Col + '1' ] = 'Replacement Value'
        self.cfgsheet[ self.VR_RVal_Col + '2' ] = '-'
        
        row = self.Config_Start_Row
        offset = 0

        #VR based action rules
        for action in self.list_of_vr_actions:
            self.cfgsheet[ self.VR_Action_Col + str(row + offset) ] = action
            offset += 1
        #flags - need to include current list of flags and default values
        self.cfgsheet[ self.FLAG_COL + '1' ] = 'FLAG'
        self.cfgsheet[ self.FLAG_VAL_COL + '1' ] = 'VALUE'

        # Note that there are _CELL definitions for each flag
        # but I am not sure how to make those without hardcoding.
        # For compatibility purposes, I am keeping few flags, and
        # keeping them in the same position
        row = self.Config_Start_Row
        offset = 0
        for flag_name in self.flag_list:
            self.cfgsheet[ self.FLAG_COL + str(row + offset)] = flag_name
            self.cfgsheet[ self.FLAG_VAL_COL + str(row + offset)] = 'FALSE'
            offset += 1

        #tags        
        self.cfgsheet[ self.Tag_Name_Col + '1' ] = 'Tag Name (autofilled)'
        self.cfgsheet[ self.Tag_Group_Col + '1' ] = 'Dicom Tag'
        self.cfgsheet[ self.Tag_Group_Col + '2' ] = 'group'
        self.cfgsheet[ self.Tag_Element_Col + '2' ] = 'element'
        self.cfgsheet[ self.Tag_Action_Col + '1' ] = 'Action'
        self.cfgsheet[ self.Tag_Action_Col + '1' ] = 'delete/replace'
        self.cfgsheet[ self.Tag_RVal_Col + '1' ] = 'Replacement Value'


    def blankTopBar( self, topbarWidth: int = 50):
        """Blanks top of [US] images to remove PHI
        Incompatible with compressed pixel data. 
        Review your images to confirm both PHI gone and images not corrupted.

        Parameters:
        topbarWidth (int): Width of strip at the top to blank
                           Deafults to 50 pixels

        Returns:
        int: 1=Successful, <0=Fail with err code
        """
        if self.DCM.file_meta.TransferSyntaxUID.is_compressed:
            # PyDICOM gets annoyed if the image data appears compressed.
            self.DCM.decompress()
        data = np.array
        data = self.DCM.pixel_array
        shape = data.shape
        data[:,1:topbarWidth,:] = 128
        self.DCM.PixelData = data.tobytes()
        return 1



        
#------------------------------------------------------------------------------------------------------


############### END OF STUDY_CLASS DEFINITION ###################


#---------------------------------------------------------
#      Accessory functions related to deidentification
#---------------------------------------------------------


def tag_data_type_callback(dataset, data_element):

    # blank all person name ('PN') values
    if data_element.VR == 'PN':
        data_element.value = 'anonymous'

    # blank all date ('DA') values
    elif data_element.VR == 'DA':
        data_element.value = ''

    # blank all time ('TM') values
    elif data_element.VR == 'TM':
        data_element.value = ''



def del_curves_callback(dataset, data_element):
    if data_element.tag.group & 0xFF00 == 0x5000:
        del dataset[data_element.tag]


#---------------------------------------------------------
#      Accessory functions related to datetime delta
#---------------------------------------------------------

def delta_apply_callback(dataset, data_element):
    '''Callback to apply datetime delta to tags.
    Used in deidentifyDICOM()
    Requires knowledge of the tag VR and delta. 
    DA = date
    TM = time
    DT = datetime
    '''
    vr = data_element.VR
    print(f'delta_apply_callback: [{data_element}], \t\t{dataset.preamble}')
    delta_obj = delta_str2obj(dataset.preamble)  # workaround to pass delta into this callback
    if vr == 'DA':
        print(f'\tdelta to DA: old= {data_element.value}', end='')
        # convert date string into date object
        old_date_obj = DA_str2obj( data_element.value )
        # add delta
        print(f' delta type={type(delta_obj)}, val={delta_obj}')
        new_date_obj = old_date_obj + delta_obj
        # convert back to string ('yyyymmdd')
        # save to original data element
        data_element.value = DA_obj2str( new_date_obj )
        print(f'\t   new= {data_element.value}')
        pass
    elif vr == 'TM':
        # convert time string into time object
        # add delta
        # convert back to string ('HHMMSS.FFFFFF')
        # save to original data element
        pass
    elif vr == 'DT':
        # convert datetime string into datetime object
        # add delta
        # convert back to string ('YYYYMMDDHHMMSS.FFFFFF') (+/- '&ZZXX' UTC info)
        # save to original data element
        pass


def DT_clean( DT_str:str )->str:
    #  Old DICOM standard allowed ':', ' ' and '.' in DA/DT/TM strings
    #  add more if necessary for non-standard formats
    DT_str = remove( DT_str, ':. ')
    return DT_str


def DA_str2obj(date_str:str)->datetime.date:
    date_str = DT_clean( date_str )
    yyyy = int(date_str[:4])
    mm = int(date_str[4:6])
    dd = int(date_str[6:8])
    date_obj = date(year=yyyy, month=mm, day=dd)
    return date_obj


def DA_obj2str(date_obj:datetime.date)->str:
    # Pad with '0' to appropriate no of digits
    print(f'DA_obj2str: in:{date_obj}')
    yyyy = str(date_obj.year).rjust(4,'0')
    mm = str(date_obj.month).rjust(2,'0')
    dd = str(date_obj.day).rjust(2,'0')
    date_str = yyyy + mm + dd
    return date_str


def TM_str2obj(time_str:str)->datetime.time:
    pass


def TM_obj2str(time_obj:datetime.time)->str:
    pass


def DT_str2obj(dt_str:str)->datetime:
    pass


def DT_obj2str(dt_obj:datetime)->str:
    pass


def delta_str2obj(delta_str:str )->timedelta:
    '''Convert string (from XLS) into timedelta object.
    expected string format 'days:seconds.miliseconds'
    Params: delta_str from XLS 'datetime delta' column
    Return: timedelta object'''
    colon = delta_str.find(':')  # should always have
    period = delta_str.find('.')  # should have. Might vary.
    # If no period? period=-1
    dd = int( delta_str[ : colon] )
    if period:
        ds = int( delta_str[colon+1 : period] )
        dms = int( delta_str[period+1 : ] )
    else:
        ds = int( delta_str[colon+1 :] )
        dms = 0
    delta_obj = timedelta( days=dd, seconds=ds, milliseconds=dms )
    return delta_obj


def delta_obj2str( delta: timedelta )->str:
    '''Take timedelta and convert into a format for storage in XLS
    eg. str_to_write = delta_as_string( time_delta obj )
    Params: timedelta object
    Return: formatted string "days:seconds.microseconds" '''
    # class datetime.timedelta(days=0, seconds=0, microseconds=0, milliseconds=0, minutes=0, hours=0, weeks=0)
    # readable properties: .days .seconds .microseconds
    days = delta.days
    seconds = delta.seconds
    raw_microseconds = delta.microseconds
    ms_string = str(raw_microseconds).zfill(6)  # force string to be 6 chars for microseconds
    delta_string = f'{days}:{seconds}.{ms_string}'
    return delta_string

######################################################################

#---------------------------------------------------------
#      Accessory functions related to creating a new XLS (mostly for Init_Study.py)
#---------------------------------------------------------


# Console text input and return valid string 
def verify_txt_input( message = 'no argument supplied'):

    inputtext = input( message )

    while not inputtext.replace(' ','').isalpha():
        print('Invalid text entry. Please enter alphabetical characters only.')
        inputtext = input( message )

    return inputtext


# Consider embedding this into creat_rnd_studyID
# to make sure the ode to create is not duplicated
# or could turn into a function instead.
def number_possible_IDs( format ):
    format = format.lower().strip()

    poss = 0
    if len(format) > 0:
        poss = 1

    for letter in format:
        if letter =='c': #if is character
            poss *= len(alphalistboth)
        elif letter =='u': #if is character
            poss *= len(alphalistupper)
        elif letter =='l': #if is character
            poss *= len(alphalistlower)
        elif letter =='d': #if is digit
            poss *= len(digitlist)

    return poss


def create_rnd_studyID( format = 'lldddd', prefix='', suffix=''):

    # c = alphabetical char, d = digit
    # u = upper, l = lower

    # c can be anything from 'a' to 'Z'
    # d can be 0-9
    # Max studyID length = 16 chars (DICOM Limit for Pt ID)
    # prefix string inserted at the beginning

    # future implementation could apply upper/lower limits 
    #  to accommodate adding to an existing list of study IDs

    # Quick validation
    
    format = format.lower()
    newID = prefix

    #1 - check input - assume input has been validated in calling function
    #                  to prevent running same checks 1000s of times

    #2 - Step through format string and append ID string with appropriate random char/digit
    letter = ''

    for letter in format.strip():
        secure_random = random.SystemRandom()
        if letter =='c': #if is character
            newID += secure_random.choice(alphalistboth)
        elif letter =='u': #if is character
            newID += secure_random.choice(alphalistupper)
        elif letter =='l': #if is character
            newID += secure_random.choice(alphalistlower)

        elif letter =='d': #if is digit
            newID += secure_random.choice(digitlist)

    newID += suffix

    return newID


def pretty_tag(group, element):
    '''Create a 'pretty' string for DICOM tag display
    input group and element hex strings
    This should match the style received from the XLS config sheet
    e.g. print( pretty_tag( "0x50", "0x8") )
    output: "(0050,0008)"
    '''
    group = str(group)[2:]
    group = group.rjust(4,"0")

    element = str(element)[2:]
    element = element.rjust(4,"0")
    
    return f'({group},{element})'


def combotag( group, element, outtype = 'int'):
    group = int(group, 16)
    element = int(element, 16)
    combotag = (group * 65536) + element
    if outtype.lower() == 'int':
        return combotag
    elif outtype.lower() == 'str':
        return hex(combotag)


def remove( inputstr:str, the_list:str)->str:
    '''Remove all occurences of certain characters from a string.
    Params: Input string(str), char list(str)
    Return: cleaned string'''
    out_str = ""
    for char in inputstr:
        if char not in the_list:
            out_str += char
    return out_str

