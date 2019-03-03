'''
StudyTools.py

Functions for use with Study anonymiser/Study ID generator

'''
import openpyxl
import random
import getpass
import os
from datetime import date, time, datetime


# Some constants to use in this library - needs to replace much of the static parts of the XLS creation:
# These are column or cell references prefixed with which sheet they belong in
xlsPage_Title              = 'A1'

xlsFront_study_title_cell    = 'B2'
xlsFront_PI_cell             = 'B3'
xlsFront_IRB_code_cell       = 'B4'
xlsFront_number_of_study_IDs_cell = 'B5' 

xlsData_study_IDs          = 'B'
xlsData_date_added         = 'C'
xlsData_time_added         = 'D'
xlsData_patient_lastname   = 'G'
xlsData_patient_firstname  = 'H'
xlsData_patient_ID         = 'I'
xlsData_accession_number   = 'K'
xlsData_study_date         = 'M'
xlsData_study_time         = 'N'
xlsData_study_UID          = 'P'
xlsData_study_description  = 'R'

xlsLog_date                = 'B'
xlsLog_time                = 'C'
xlsLog_activity            = 'E'
xlsLog_user                = 'G'
xlsLog_computer            = 'H'

xls_UID_lookup = {}

alphalistboth = ['a','b','c','d','e','f','g','h','i','j','k','l','m','o','n','p','q','r',
                 's','t','u','v','w','x','y','z','A','B','C','D','E','F','G','H','I','J',
                 'K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
alphalistupper = ['A','B','C','D','E','F','G','H','I','J',
                  'K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
alphalistlower = ['a','b','c','d','e','f','g','h','i','j','k','l','m','o','n','p','q','r',
                  's','t','u','v','w','x','y','z' ]

digitlist = [ '0','1','2','3','4','5','6','7','8','9' ]


# Console text input and return valid string 
def verify_txt_input( message = 'no argument supplied'):

	inputtext = input( message )

	while not inputtext.replace(' ','').isalpha():
		print('Invalid text entry. Please enter alphabetical characters only.')
		inputtext = input( message )

	return inputtext


def number_possible_IDs( format ):
	format = format.lower().strip()

	poss = 0
	if len(format) > 0:
		poss = 1

	for letter in format:
		if letter =='c': #if is character
			poss *= len(alphalistboth)
		elif letter =='u': #if is character
			poss *= len(alphalistupper)
		elif letter =='l': #if is character
			poss *= len(alphalistlower)
		elif letter =='d': #if is digit
			poss *= len(digitlist)

	return poss




def create_rnd_studyID( format = 'lldddd', prefix=''):

	# c = alphabetical char, d = digit

	# c can be anything from 'a' to 'Z'
	# d can be 0-9
	# Max studyID length = 16 chars (DICOM Limit for Pt ID)
	# prefix string inserted at the beginning

	# future implementation could apply upper/lower limits 
	#  to accommodate adding to an existing list of study IDs

	# Quick validation
	
	format = format.lower()
	newID = prefix

	#1 - check input - assume input has been validated in calling function
	#                  to prevent running same checks 1000s of times

	#2 - Step through format string and append ID string with appropriate random char/digit
	letter = ''

	for letter in format.strip():
		secure_random = random.SystemRandom()
		if letter =='c': #if is character
			newID += secure_random.choice(alphalistboth)
		elif letter =='u': #if is character
			newID += secure_random.choice(alphalistupper)
		elif letter =='l': #if is character
			newID += secure_random.choice(alphalistlower)

		elif letter =='d': #if is digit
			newID += secure_random.choice(digitlist)

	return newID


# ---------------------> Moved from Init_Study
# should return the openpyxl workbook object 

def create_new_study_xls_file( new_xlsfilename,
                               new_study_title,
	    					   new_primary_investigator,
							   new_number_of_study_IDs ):


	new_workbook = openpyxl.Workbook()

	try:
		new_workbook.save( new_xlsfilename )
	except:
		print(f'Fatal Error: Failed to save \"{ new_xlsfilename }\"')
		raise
	else:
		print(f'Created blank template file \"{ new_xlsfilename }\" OK')


	new_wsFront       = new_workbook.active
	new_wsFront.title = 'Front'

	new_wsData = new_workbook.create_sheet('Data') # insert log sheet at the end
	new_wsLog  = new_workbook.create_sheet('Log' ) # insert log sheet at the end

	#--------------------->  Add in basic data  <-------------------------

	# Col A is the list of titles. Text info is in col B.
	new_wsFront['A1'] = 'Front Page'            
	new_wsFront['A' + str( xlsFront_study_title_cell[-1] ) ]         = 'Study Title:'          # data held in xlsFront_study_title_cell
	new_wsFront['A' + str( xlsFront_PI_cell[-1] )          ]         = 'Primary Investigator:' # data held in xlsFront_PI_cell
	new_wsFront['A' + str( xlsFront_number_of_study_IDs_cell[-1] ) ] = 'No of Study IDs'       # data held in xlsFront_number_of_study_IDs_cell

	new_wsFront[ xlsFront_study_title_cell ]        = new_study_title
	new_wsFront[ xlsFront_PI_cell ]                 = new_primary_investigator
	new_wsFront[ xlsFront_number_of_study_IDs_cell ] = new_number_of_study_IDs

	# Row 1 is the column title row
	new_wsData['A1']  = 'Data Page'
	new_wsData[ xlsData_study_IDs  + '1']  = 'Study IDs'
	new_wsData[ xlsData_date_added + '1']  = 'Date Added'
	new_wsData[ xlsData_time_added + '1']  = 'Time Added'

	new_wsData[ xlsData_patient_lastname  + '1'] = 'Patient Last Name'
	new_wsData[ xlsData_patient_firstname + '1'] = 'First Name'

	new_wsData[ xlsData_patient_ID       + '1']  = 'Patient ID'
	new_wsData[ xlsData_accession_number + '1']  = 'Accession No.'
	new_wsData[ xlsData_study_date       + '1']  = 'Study Date'
	new_wsData[ xlsData_study_time       + '1']  = 'Study Time'
	new_wsData[ xlsData_study_UID        + '1']  = 'Study UID'
	new_wsData[ xlsData_study_description + '1'] = 'Study Description'

	new_wsLog['A1']   = 'Log Page'

	new_wsLog[ xlsLog_date + '1']   = 'Date'
	new_wsLog[ xlsLog_time + '1']   = 'Time'
	new_wsLog[ xlsLog_activity + '1']   = 'Log Activity'
	new_wsLog[ xlsLog_user + '1']   = 'User'
	new_wsLog[ xlsLog_computer + '1']   = 'Computer'

	return new_workbook
#------------------------------------------------------------------------------------------------------
def load_study_xls( xls_filename ):
	"""
	Loads study XLS file, and does basic checking to make sure that it is actually a valid XLS sheet with check_xls_is_valid().
	Usage: study_workbook = load_study_xls( <filename> )
	Returns: openpyxl XLS workbook object
	"""

	new_xls_UID_lookup = {}

	try:
		new_workbook = openpyxl.load_workbook( xls_filename )
	except:
		return False, False

	if check_xls_is_valid( new_workbook ):
		new_xls_UID_lookup = cache_existing_xls_UIDs( new_workbook )
		return new_workbook, new_xls_UID_lookup

	else:
		return False, False




#------------------------------------------------------------------------------------------------------
def check_xls_is_valid( xls_workbook ):
	"""check_xls_is_valid() is a simple check to see if the expected Front/Data/Log worksheet structure is present.
	This could become more elaborate over time- additional options to check most recent addition etc.
	Usage:  check_xls_is_valid( <openpyxl workbook object> )
	Returns: True (if valid), or False (if considered invalid) 	"""

	if xls_workbook.sheetnames != ['Front', 'Data', 'Log']:
		# print(f'Fatal Error: {xls_filename} is not a valid DeIdentifier XLSX file.')
		return False
	else:
		# print(f'Loaded {xls_filename} OK')
		return True


#------------------------------------------------------------------------------------------------------

def cache_existing_xls_UIDs ( wb ):
	"""
	Identifies each stored study UID and caches it and the corresponding row number in a dict
	Returns the absolute row number - not relative, so starts at 2 (as does the data) and ends at total_studyIDs + 1
	Usage: dict_cache = first_blank_ptID_row =  cache_existing_xls_UIDs ( <openpyxl xls workbook> )
	Returns: dict storing UIDs as Keys, and corresponding rows as Values. 
	"""
	wsData = wb['Data']
	wsFront = wb['Front']

	dict_cache = {}

	row = 2
	max_row = int( wsFront[ xlsFront_number_of_study_IDs_cell ].value ) + 1  # +1 as the data-containing rows start at row 2, not 1

	check_ptID      = wsData[ xlsData_patient_ID + str(row)].value
	check_studyID   = wsData[ xlsData_study_IDs  + str(row)].value
	check_study_UID = wsData[ xlsData_study_UID  + str(row)].value

	while (check_studyID != '') and (check_ptID != '') and (row <= max_row ) and (check_study_UID !=''):
		dict_cache[ check_study_UID ] = row
		row += 1
		check_ptID      = wsData[ xlsData_patient_ID + str(row)].value
		check_studyID   = wsData[ xlsData_study_IDs  + str(row)].value
		check_study_UID = wsData[ xlsData_study_UID  + str(row)].value

	return dict_cache



#------------------------------------------------------------------------------------------------------


def get_DCM_StudyInstanceUID( dicom_object ):
	'''
	Returns the Study Instance UID from the specified DICOM file object
	Usage:   string_variable = get_DCM_StudyInstanceUID ( <dicom_object> )
	Returns:  a string containing the study instance UID.
	          returns FALSE if no StudyInstanceUID tag is found.
	This iterates through all tags until it finds the correct one.
	  This bypasses the issue with DICOMDIR hiding it in a ?series?
	  which is not visible to the standard dicom_object.StudyInstanceUID method.
	'''
	siUID = False

	for elem in dicom_object.iterall():
		if 'Study Instance UID' == elem.name:
			siUID = elem.value
			break

	return siUID



#------------------------------------------------------------------------------------------------------


def first_available_studyID_row ( wb ):
	"""
	Identifies the FIRST available studyID in the XLS file
	This is slow, so separated from (but called by) the NEXT study ID generator
	Returns the absolute row number - not relative, so starts at 2 (as does the data) and ends at total_studyIDs + 1
	Usage: first_blank_ptID_row =  first_available_studyID_row ( <openpyxl xls workbook> )
	Returns: 1st free row where a studyID can be assigned. 
	"""
	wsData = wb['Data']
	wsFront = wb['Front']


	row = 2
	max_row = wsFront['B4'].value + 1  # +1 as the data-containing rows start at row 2, not 1

	check_ptID    = wsData[ xlsData_patient_ID + str(row)].value
	check_studyID = wsData[ xlsData_study_IDs  + str(row)].value
 
	while (check_studyID != None ) and (check_ptID != None ) and (row <= max_row ):
		row += 1
		check_ptID    = wsData[ xlsData_patient_ID + str(row)].value
		check_studyID = wsData[ xlsData_study_IDs  + str(row)].value

	return row


#------------------------------------------------------------------------------------------------------

def next_XLSrow_gen ( start_row, 
                            n_data,
							worksheet,
							column ):
	'''
	Generator to return the cell value of a specified column in the XLS
	eg. for StudyID:
	next_study_ID_generator = next_XLSrow_gen( startrow, totalStudyIDs, wsData, 'B' )
	Returns: generator object
	'''

	for row in range( start_row, (n_data + 2) ):
		data = worksheet[ column + str( row ) ].value
		
		if data != None:
			yield data, row
		else:
			yield False, row


#------------------------------------------------------------------------------------------------------

# function that takes raw identifying data, 
# - identifies the next available studyID in the worksheet
# - inserts this data into the XLS
# - returns the studyID to be given to the deidentifier


def assign_next_free_studyID ( wb,   #XLS openpyxl workbook object
                               dicom ):  #pyDICOM object


	wsData = wb['Data']
	wsFront = wb['Front']

	starting_row = first_available_studyID_row( wb )
	no_of_studyIDs  = wsFront[ xlsFront_number_of_study_IDs_cell ].value

	#generator logic here to move to next row each call
	# and insert new study details into XLS sheet

	next_study_ID_generator = next_XLSrow_gen( starting_row, no_of_studyIDs, wsData, xlsData_study_IDs )

	for studyID, current_row in next_study_ID_generator:   # last_row +1 to include 'last_row' within the for loop
		if studyID:
			# ie if the generator has successfully returned something. It has not been checked for validity

			# insert Logging message(s) here
			
			wsData[ xlsData_patient_ID        + str( current_row ) ] = str( dicom.PatientID )
			wsData[ xlsData_patient_lastname  + str( current_row ) ] = str( dicom.PatientName )
		
			wsData[ xlsData_accession_number  + str( current_row ) ] = str( dicom.AccessionNumber )
			wsData[ xlsData_study_date        + str( current_row ) ] = str( dicom.StudyDate )
			wsData[ xlsData_study_time        + str( current_row ) ] = str( dicom.StudyTime )
			wsData[ xlsData_study_UID         + str( current_row ) ] = str( dicom.StudyInstanceUID )
			wsData[ xlsData_study_description + str( current_row ) ] = str( dicom.StudyDescription )

			# Yield the new studyID string  
			return studyID
		else:
			#This will happen if we have run out of possible studyIDs
			# Actually, this 'else' statement is redundant but makes the code clearer to me.
			return False


#------------------------------------------------------------------------------------------------------

def log_xls_creation ( workbook,
                       no_of_studyIDs,
					   studyID_prefix,
					   studyID_format ):

	wsLog   = workbook[ 'Log' ]

	username   = getpass.getuser()
	compname   = os.environ['COMPUTERNAME']
	dateobject = datetime.now()
	date       = dateobject.strftime('%d-%m-%Y')
	timenow    = dateobject.strftime('%S:%M:%H')

	row_text = '2'  # This can be set firmly as it is a new XLS file

	wsLog[ xlsLog_date     + row_text ] =  date  #'Date'
	wsLog[ xlsLog_time     + row_text ] =  timenow  #'Time'
	wsLog[ xlsLog_activity + row_text ] =  f'Created XLSX file (n={ no_of_studyIDs }, format={ studyID_prefix }{ studyID_format })'  #'Log Activity'
	wsLog[ xlsLog_user     + row_text ] =  username  #'User'
	wsLog[ xlsLog_computer + row_text ] =  compname  #'Computer'



#------------------------------------------------------------------------------------------------------

def write_xls_to_disc ( workbook, filename ):

	workbook.save( filename )


#------------------------------------------------------------------------------------------------------



def get_old_study_attrib( workbook, dict_cache, studyUID, attribute ):
	'''
	Returns string from relevent cell in Data sheet from the workbook.
	Takes the studyUID and str 'attribute' as reference
	Usage: get_old_study_attrib( <openpyxl workbook>, <existing study UID>, <str attribute> )
	'attribute' should be txt indicating the relevent data page column. Use the static xls_Data_... column values from studytools.py
	eg. xls_Data_study_UID
	'''
	wsData = workbook[ 'Data' ]

	old_study_ID = wsData[ attribute + str( dict_cache[ studyUID ] )  ].value

	return old_study_ID

